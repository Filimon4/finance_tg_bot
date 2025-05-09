from asyncio.log import logger
from datetime import datetime, timedelta
import re
import tempfile
from openpyxl import Workbook, load_workbook
from typing import Generator
import os
from src.db.index import DB
from src.modules.finance.operations.operationsRepository import OperationsRepository
from sqlalchemy.orm import Session
from dataclasses import dataclass
from aiogram.types import Message, BufferedInputFile
from sqlalchemy.exc import SQLAlchemyError

from src.telegramBot.bot.BotTg import MainBotTg

@dataclass
class ReportDTO:
    date: str
    category: str
    amount: float
    description: str
    cash_account: str
    name: str
    user_id: int

class ExcelReportGenerator:

    @staticmethod
    async def on_command_export(user_id: int, text: str):
        try:
            matched = re.match(r'^(\/export)\s+(\d+)$', text)
            if not matched:
                await MainBotTg.send_message(chat_id=user_id, text='Неправильный шаблон команды. Пример использования: /export 12')
                return

            _, month = matched.groups()
            month = int(month)
            await MainBotTg.send_message(chat_id=user_id, text=f'🔍 Формирую отчет за {month} месяцев...')

            with DB.get_session() as session:
                await ExcelReportGenerator.generate_and_send_report(
                    month=month,
                    user_id=user_id,
                    session=session
                )

        except SQLAlchemyError as e:
            logger.error(f"Database error: {str(e)}")
            await MainBotTg.answer(text="❌ Ошибка базы данных при экспорте")
        except Exception as e:
            logger.error(f"Export error: {str(e)}")
            await MainBotTg.answer(text="❌ Произошла ошибка при формировании отчета")

    @staticmethod
    async def generate_and_send_report(
        month: int,
        user_id: int,
        session: Session
    ) -> None:
        """Статический метод для генерации и отправки отчета"""
        try:
            end_time = datetime.now()
            start_time = end_time - timedelta(days=30*month)

            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
                report_path = tmp_file.name

            try:
                report_generator = ExcelReportGenerator(report_path)
                report_generator.generate_operations_report(
                    session=session,
                    user_id=user_id,
                    start_time=start_time,
                    end_time=end_time
                )

                with open(report_path, 'rb') as f:
                    file_data = f.read()
                
                await MainBotTg.send_document(
                    chat_id=user_id,
                    document=BufferedInputFile(
                        file=file_data,
                        filename=f"financial_report_{month}_months.xlsx"
                    ),
                    caption=f"📊 Финансовый отчет за {month} месяцев"
                )

            finally:
                # Гарантированно удаляем временный файл
                if os.path.exists(report_path):
                    os.remove(report_path)

        except Exception as e:
            logger.error(f"Error in generate_and_send_report: {str(e)}")
            raise


    def __init__(self, report_file: str = "operations_report.xlsx"):
        self.report_file = report_file
        self.use_tempfile = False
        self._temp_file_created = False

    def generate_operations_report(self, session: Session, user_id: int, start_time: datetime, end_time: datetime, page_size: int = 40) -> str:
        try:
            self._initialize_workbook()
            
            wb = load_workbook(self.report_file)
            ws = wb["data"]
            
            for operations in self.paginatedOperationsGenerator(session, user_id, start_time, end_time, page_size):
                for operation in operations:
                    report_data = ReportDTO(
                        date=operation.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                        category=operation.category.name if operation.category else "",
                        amount=float(operation.amount),
                        description=operation.description or "",
                        user_id=operation.account_id or 0,
                        name=operation.name,
                        cash_account=operation.cash_account.name
                    )
                    next_row = ws.max_row + 1
                    data_values = [
                        report_data.date,
                        report_data.category,
                        report_data.cash_account,
                        report_data.amount,
                        report_data.name,
                        report_data.description,
                    ]
                    
                    for col_num, value in enumerate(data_values, start=1):
                        ws.cell(row=next_row, column=col_num, value=value)
            
            wb.save(self.report_file)
            wb.close()
            return self.report_file

        except Exception as e:
            raise e

    #region Temp file

    def _initialize_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "data"
        headers = ["Дата", "Категория", "Счёт", "Сумма", "Название", "Описание"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        wb.save(self.report_file)
        wb.close()

    #region Generator

    def paginatedOperationsGenerator(self, session: Session, user_id: int, start_time: datetime, end_time: datetime, page_size: int = 40) -> Generator:
        page = 1
        while True:
            data_operations = OperationsRepository.getPaginatedOperations(session, user_id, start_time, end_time, page, page_size)
            if not data_operations['operations']:
                break
                
            yield data_operations['operations']
            page += 1